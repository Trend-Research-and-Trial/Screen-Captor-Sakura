import openpyxl
from openpyxl.styles import Font
import math
from settings import OUTPUT

class Excel:
    """
    Args:
        path str:
        row int:
        column int:
        mode int: 0 => 新規作成モード, 1 => 追記モード
    """
    def __init__(self, path, row=1, column=1, cell=None, mode=0):

        self.path = path

        if mode==0:
            # 0の場合：新規作成モード
            # Excelファイルの新規作成
            self.book = openpyxl.Workbook()
            self.__initializeCellPosition(row, column, cell)

            sheet, current_cell = self.__getSheetAndCell()

            sheet.title = "手順書" 

            # 指定したセルにタイトルを入力する
            current_cell.value = self.path.split("/")[-1].split(".")[0]
            
            self.next_row = row + 2
            self.next_column = column

        elif mode == 1:
            # 1の場合：追記モード
            self.book = openpyxl.load_workbook(self.path)
            self.__initializeCellPosition(row, column, cell)

        self.book.save(self.path)

    def paste(self, value):
        # シートの取得とシート名の変更
        _, current_cell = self.__getSheetAndCell()

        # 指定したセルに文字列を入力する
        current_cell.value = value
        
        self.__save(row_to_add=2)
        
    def pasteImg(self, value, width, hight):
        # シートの取得とシート名の変更
        sheet, current_cell = self.__getSheetAndCell()
        
        img = openpyxl.drawing.image.Image(value)
        str_current_cell= current_cell.coordinate  # カレントのセルを'A1'のような形式に変換(次行の処理のため)
        img.anchor = str_current_cell
        sheet.add_image(img)

        self.__save(row_to_add=math.ceil(hight*1.25 / 22 + 1))
    
    def __save(self, row_to_add=0, column_to_add=0):
        self.next_row += row_to_add
        self.next_column += column_to_add
        self.book.save(self.path)

    def __getSheetAndCell(self):
        sheet = self.book.active
        current_cell = sheet.cell(self.next_row, self.next_column)
        return sheet, current_cell
    
    def __initializeCellPosition(self, row, column, cell):
        self.next_row = row
        self.next_column = column
        if cell != None:
            sheet, _ = self.__getSheetAndCell()
            self.next_row = sheet[cell].row
            self.next_column = sheet[cell].column